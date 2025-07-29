from fastapi import HTTPException
from sqlalchemy.orm import Session, joinedload
from models.models import Producto
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

def get_productos(db: Session):
    # Incluir las relaciones marca y categoria al hacer la consulta
    productos = db.query(Producto).options(
        joinedload(Producto.marca),
        joinedload(Producto.categoria)
    ).all()
    
    # Procesar las imágenes para tener URLs completas
    for producto in productos:
        if producto.imagen and not producto.imagen.startswith('http'):
            # Si la imagen no empieza con http, construir la URL completa
            if producto.imagen.startswith('/uploads/'):
                producto.imagen = f"http://127.0.0.1:8000{producto.imagen}"
            elif not producto.imagen.startswith('/'):
                producto.imagen = f"http://127.0.0.1:8000/uploads/productos/{producto.imagen}"
    
    return productos

def get_producto(db: Session, id_producto: int):
    producto = db.query(Producto).options(
        joinedload(Producto.marca),
        joinedload(Producto.categoria)
    ).filter(Producto.id_producto == id_producto).first()
    
    if not producto:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    
    # Procesar la imagen para tener URL completa
    if producto.imagen and not producto.imagen.startswith('http'):
        if producto.imagen.startswith('/uploads/'):
            producto.imagen = f"http://127.0.0.1:8000{producto.imagen}"
        elif not producto.imagen.startswith('/'):
            producto.imagen = f"http://127.0.0.1:8000/uploads/productos/{producto.imagen}"
    
    return producto

def create_producto(db: Session, producto_data: dict):
    # Validar campos requeridos
    required_fields = ['nombre', 'id_marca', 'stock', 'precio_mayorista', 'precio_minorista', 'id_categoria']
    for field in required_fields:
        if field not in producto_data:
            raise HTTPException(status_code=422, detail=f"Campo requerido faltante: {field}")
    
    try:
        nuevo_producto = Producto(
            nombre=producto_data['nombre'],
            id_marca=producto_data['id_marca'],
            stock=producto_data['stock'],
            precio_mayorista=producto_data['precio_mayorista'],
            precio_minorista=producto_data['precio_minorista'],
            id_categoria=producto_data['id_categoria'],
            iva=producto_data.get('iva', 0.12),
            estado=producto_data.get('estado', 'activo'),
            imagen=producto_data.get('imagen')
        )
        
        db.add(nuevo_producto)
        db.commit()
        db.refresh(nuevo_producto)
        
        # Cargar las relaciones después del commit
        db.refresh(nuevo_producto)
        nuevo_producto = db.query(Producto).options(
            joinedload(Producto.marca),
            joinedload(Producto.categoria)
        ).filter(Producto.id_producto == nuevo_producto.id_producto).first()
        
        return nuevo_producto
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=422, detail=str(e))

def update_producto(db: Session, id_producto: int, producto_data: dict):
    producto = db.query(Producto).filter(Producto.id_producto == id_producto).first()
    if not producto:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    
    for key, value in producto_data.items():
        setattr(producto, key, value)
    
    db.commit()
    db.refresh(producto)
    
    # Cargar las relaciones después del commit
    producto = db.query(Producto).options(
        joinedload(Producto.marca),
        joinedload(Producto.categoria)
    ).filter(Producto.id_producto == id_producto).first()
    
    return producto

def delete_producto(db: Session, id_producto: int):
    producto = db.query(Producto).filter(Producto.id_producto == id_producto).first()
    if not producto:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    
    # Eliminar archivo de imagen si existe
    if producto.imagen:
        import os
        try:
            # Construir la ruta del archivo
            if producto.imagen.startswith('/uploads/'):
                file_path = producto.imagen[1:]  # Remover el primer '/'
            else:
                file_path = f"uploads/productos/{producto.imagen}"
            
            if os.path.exists(file_path):
                os.remove(file_path)
        except Exception as e:
            print(f"Error al eliminar imagen: {e}")
    
    db.delete(producto)
    db.commit()
    return {"mensaje": "Producto eliminado"}

def export_productos_to_excel(db: Session):
    """
    Exporta los productos a un archivo Excel con diseño de tabla
    """
    try:
        productos = db.query(Producto).options(
            joinedload(Producto.marca),
            joinedload(Producto.categoria)
        ).all()
        
        if not productos:
            raise HTTPException(status_code=404, detail="No hay productos para exportar")
        
        # Preparar datos para el DataFrame
        data = []
        for producto in productos:
            data.append({
                'ID Producto': producto.id_producto or '',
                'Nombre': producto.nombre or '',
                'Marca': producto.marca.descripcion if producto.marca else 'Sin marca',
                'Categoría': producto.categoria.descripcion if producto.categoria else 'Sin categoría',
                'Stock': producto.stock or 0,
                'Precio Minorista': f"${producto.precio_minorista:.2f}" if producto.precio_minorista else "$0.00",
                'Precio Mayorista': f"${producto.precio_mayorista:.2f}" if producto.precio_mayorista else "$0.00",
                'IVA': f"{(producto.iva * 100):.1f}%" if producto.iva else "0.0%",
                'Estado': producto.estado.capitalize() if producto.estado else 'Inactivo',
                'Tiene Imagen': 'Sí' if producto.imagen else 'No'
            })
        
        # Crear DataFrame
        df = pd.DataFrame(data)
        
        # Crear archivo Excel con diseño profesional
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventario de Productos"
        
        # Configurar estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")  # Verde como en el formulario
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Estilo para las celdas de datos
        data_alignment = Alignment(horizontal="left", vertical="center")
        price_alignment = Alignment(horizontal="right", vertical="center")  # Para precios alineados a la derecha
        center_alignment = Alignment(horizontal="center", vertical="center")  # Para stock y estado
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Agregar título principal
        ws.merge_cells('A1:J1')  # 10 columnas para productos
        title_cell = ws['A1']
        title_cell.value = f"INVENTARIO DE PRODUCTOS - {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")  # Verde claro
        
        # Agregar una fila vacía
        ws.append([])
        
        # Agregar encabezados
        headers = list(df.columns)
        ws.append(headers)
        
        # Aplicar estilo a los encabezados
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Agregar datos
        for row_num, row_data in enumerate(df.values, 4):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = thin_border
                
                # Aplicar alineación específica según el tipo de dato
                if col_num in [6, 7]:  # Precios (columnas 6 y 7)
                    cell.alignment = price_alignment
                elif col_num in [1, 5, 9, 10]:  # ID, Stock, Estado, Tiene Imagen
                    cell.alignment = center_alignment
                else:
                    cell.alignment = data_alignment
                
                # Aplicar color especial para stock bajo
                if col_num == 5:  # Columna de stock
                    try:
                        stock_value = int(str(value))
                        if stock_value <= 10:
                            cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")  # Rojo claro
                        elif stock_value <= 20:
                            cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")  # Naranja claro
                    except:
                        pass
                
                # Aplicar color para estado
                if col_num == 9:  # Columna de estado
                    if str(value).lower() == 'activo':
                        cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")  # Verde claro
                    else:
                        cell.fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")  # Rojo claro
                
                # Aplicar color alternado a las filas (solo si no hay color específico)
                elif row_num % 2 == 0 and col_num not in [5, 9]:
                    cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        
        # Ajustar ancho de columnas
        column_widths = {
            'A': 12,  # ID Producto
            'B': 30,  # Nombre
            'C': 20,  # Marca
            'D': 20,  # Categoría
            'E': 10,  # Stock
            'F': 15,  # Precio Minorista
            'G': 15,  # Precio Mayorista
            'H': 10,  # IVA
            'I': 12,  # Estado
            'J': 12   # Tiene Imagen
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Agregar información de resumen al final
        last_row = len(df) + 5
        
        # Total de productos
        ws.cell(row=last_row, column=1, value="Total de productos:").font = Font(bold=True)
        ws.cell(row=last_row, column=2, value=len(df)).font = Font(bold=True)
        
        # Productos activos
        productos_activos = len([p for p in productos if p.estado == 'activo'])
        ws.cell(row=last_row + 1, column=1, value="Productos activos:").font = Font(bold=True)
        ws.cell(row=last_row + 1, column=2, value=productos_activos).font = Font(bold=True)
        
        # Productos con stock bajo (<=10)
        productos_stock_bajo = len([p for p in productos if p.stock and int(str(p.stock)) <= 10])
        ws.cell(row=last_row + 2, column=1, value="Productos con stock bajo (≤10):").font = Font(bold=True)
        ws.cell(row=last_row + 2, column=2, value=productos_stock_bajo).font = Font(bold=True, color="FF0000")
        
        # Valor total del inventario (precio minorista * stock)
        valor_total = sum([
            (p.precio_minorista or 0) * (int(str(p.stock)) if p.stock and str(p.stock).isdigit() else 0)
            for p in productos
        ])
        ws.cell(row=last_row + 3, column=1, value="Valor total inventario (minorista):").font = Font(bold=True)
        ws.cell(row=last_row + 3, column=2, value=f"${valor_total:.2f}").font = Font(bold=True, color="2E7D32")
        
        # Crear el archivo en memoria
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Retornar los bytes del archivo
        return excel_buffer.getvalue()
        
    except Exception as e:
        print(f"Error detallado al exportar productos a Excel: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error al generar archivo Excel: {str(e)}")