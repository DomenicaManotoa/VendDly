from fastapi import HTTPException
from sqlalchemy.orm import Session
from models.models import Usuario, Rol
from utils.security import hash_password_with_salt, generate_salt
# ← Nuevas importaciones para Excel
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

def get_usuarios(db: Session):
    """
    Obtiene todos los usuarios con sus roles incluidos
    """
    try:
        # Usar joinedload para cargar la relación rol
        from sqlalchemy.orm import joinedload
        usuarios = db.query(Usuario).options(joinedload(Usuario.rol)).all()
        return usuarios
    except Exception as e:
        print(f"Error al obtener usuarios: {e}")
        raise HTTPException(status_code=500, detail="Error interno del servidor")

def get_usuario(db: Session, identificacion: str):
    """
    Obtiene un usuario específico con su rol incluido
    """
    try:
        from sqlalchemy.orm import joinedload
        usuario = db.query(Usuario).options(joinedload(Usuario.rol)).filter(Usuario.identificacion == identificacion).first()
        if not usuario:
            raise HTTPException(status_code=404, detail="Usuario no encontrado")
        return usuario
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error al obtener usuario: {e}")
        raise HTTPException(status_code=500, detail="Error interno del servidor")

def create_usuario(db: Session, usuario_data: dict):
    """
    Crea un nuevo usuario con salt
    """
    try:
        # CAMBIO 1: Generar salt y hashear la contraseña antes de guardar
        if 'contrasena' in usuario_data and usuario_data['contrasena']:
            salt = generate_salt()
            usuario_data['salt'] = salt
            usuario_data['contrasena'] = hash_password_with_salt(usuario_data['contrasena'], salt)
            print(f"Usuario creado - Salt generado y contraseña hasheada")
        
        nuevo_usuario = Usuario(**usuario_data)
        db.add(nuevo_usuario)
        db.commit()
        db.refresh(nuevo_usuario)
        
        # Cargar el rol después de crear
        from sqlalchemy.orm import joinedload
        usuario_con_rol = db.query(Usuario).options(joinedload(Usuario.rol)).filter(Usuario.identificacion == nuevo_usuario.identificacion).first()
        return usuario_con_rol
    except Exception as e:
        db.rollback()
        print(f"Error al crear usuario: {e}")
        raise HTTPException(status_code=500, detail="Error al crear usuario")

def update_usuario(db: Session, identificacion: str, usuario_data: dict):
    """
    Actualiza un usuario existente con salt - CORREGIDO para evitar doble hash
    """
    try:
        usuario = db.query(Usuario).filter(Usuario.identificacion == identificacion).first()
        if not usuario:
            raise HTTPException(status_code=404, detail="Usuario no encontrado")
        
        # CAMBIO 2: Solo hashear si realmente hay una nueva contraseña
        if 'contrasena' in usuario_data:
            if usuario_data['contrasena'] and usuario_data['contrasena'].strip():
                # Solo hashear si hay una contraseña nueva no vacía
                salt = generate_salt()
                usuario_data['salt'] = salt
                usuario_data['contrasena'] = hash_password_with_salt(usuario_data['contrasena'], salt)
                print(f"Usuario editado - Nueva contraseña hasheada con nuevo salt")
            else:
                # Si la contraseña está vacía o solo espacios, no actualizarla
                print(f"Usuario editado - Manteniendo contraseña existente")
                del usuario_data['contrasena']
                # CAMBIO 3: No tocar el salt existente si no se cambia la contraseña
                if 'salt' in usuario_data:
                    del usuario_data['salt']
        
        # CAMBIO 4: Actualizar solo los campos que realmente cambiaron
        for key, value in usuario_data.items():
            if hasattr(usuario, key):  # Verificar que el atributo existe
                setattr(usuario, key, value)
                print(f"Actualizando campo {key}")
        
        db.commit()
        db.refresh(usuario)
        
        # Cargar el rol después de actualizar
        from sqlalchemy.orm import joinedload
        usuario_con_rol = db.query(Usuario).options(joinedload(Usuario.rol)).filter(Usuario.identificacion == identificacion).first()
        return usuario_con_rol
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"Error al actualizar usuario: {e}")
        raise HTTPException(status_code=500, detail="Error al actualizar usuario")

def delete_usuario(db: Session, identificacion: str):
    """
    Elimina un usuario
    """
    try:
        usuario = db.query(Usuario).filter(Usuario.identificacion == identificacion).first()
        if not usuario:
            raise HTTPException(status_code=404, detail="Usuario no encontrado")
        
        db.delete(usuario)
        db.commit()
        return {"mensaje": "Usuario eliminado correctamente"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"Error al eliminar usuario: {e}")
        raise HTTPException(status_code=500, detail="Error al eliminar usuario")

# ← Nueva función para exportar a Excel
def export_usuarios_to_excel(db: Session):
    """
    Exporta los usuarios a un archivo Excel con diseño de tabla
    """
    try:
        # Obtener usuarios con sus roles
        from sqlalchemy.orm import joinedload
        usuarios = db.query(Usuario).options(joinedload(Usuario.rol)).all()
        
        if not usuarios:
            raise HTTPException(status_code=404, detail="No hay usuarios para exportar")
        
        # Preparar datos para el DataFrame
        data = []
        for usuario in usuarios:
            # Obtener descripción del rol
            rol_descripcion = "Sin rol"
            if usuario.rol:
                rol_descripcion = usuario.rol.descripcion
            
            # Formatear fecha
            fecha_str = ""
            if usuario.fecha_actualizacion:
                fecha_str = usuario.fecha_actualizacion.strftime("%d/%m/%Y")
            
            data.append({
                'Identificación': usuario.identificacion,
                'RUC Empresarial': usuario.rucempresarial or '',
                'Nombre': usuario.nombre,
                'Correo': usuario.correo,
                'Celular': usuario.celular,
                'Estado': usuario.estado,
                'Rol': rol_descripcion,
                'Fecha Actualización': fecha_str
            })
        
        # Crear DataFrame
        df = pd.DataFrame(data)
        
        # Crear archivo Excel con diseño profesional
        wb = Workbook()
        ws = wb.active
        ws.title = "Lista de Usuarios"
        
        # Configurar estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Estilo para las celdas de datos
        data_alignment = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Agregar título principal
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = f"LISTA DE USUARIOS - {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Agregar una fila vacía
        ws.append([])
        
        # Agregar encabezados
        headers = list(df.columns)
        ws.append(headers)
        
        # Aplicar estilo a los encabezados
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Agregar datos
        for row_num, row_data in enumerate(df.values, 4):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = data_alignment
                cell.border = thin_border
                
                # Aplicar color alternado a las filas
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                
                # Aplicar formato especial a la columna Estado
                if col_num == 6:  # Columna Estado
                    if value == 'activo':
                        cell.font = Font(color="008000", bold=True)  # Verde
                    elif value == 'inactivo':
                        cell.font = Font(color="FF0000", bold=True)  # Rojo
        
        # Ajustar ancho de columnas
        column_widths = {
            'A': 15,  # Identificación
            'B': 20,  # RUC Empresarial
            'C': 25,  # Nombre
            'D': 30,  # Correo
            'E': 15,  # Celular
            'F': 12,  # Estado
            'G': 20,  # Rol
            'H': 18   # Fecha Actualización
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Agregar información de resumen al final
        last_row = len(df) + 5
        ws.cell(row=last_row, column=1, value="Total de usuarios:").font = Font(bold=True)
        ws.cell(row=last_row, column=2, value=len(df)).font = Font(bold=True)
        
        # Crear el archivo en memoria
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer.getvalue()
        
    except Exception as e:
        print(f"Error al exportar usuarios a Excel: {e}")
        raise HTTPException(status_code=500, detail="Error al generar archivo Excel")