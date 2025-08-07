from fastapi import HTTPException
from sqlalchemy.orm import Session, joinedload
from models.models import Cliente, UbicacionCliente
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime


def get_clientes(db: Session):
    """
    Obtiene todos los clientes con información de su ubicación principal
    """
    return db.query(Cliente).options(
        joinedload(Cliente.ubicaciones)
    ).all()

def get_cliente(db: Session, cod_cliente: str):
    """
    Obtiene un cliente específico con información de su ubicación principal
    """
    cliente = db.query(Cliente).options(
        joinedload(Cliente.ubicaciones)
    ).filter(Cliente.cod_cliente == cod_cliente).first()
    
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    return cliente

def create_cliente(db: Session, cliente_data: dict):
    """
    Crea un nuevo cliente con validación de ubicación principal
    """
    try:
        # Validar ubicación principal si se proporciona
        if cliente_data.get('id_ubicacion_principal'):
            ubicacion = db.query(UbicacionCliente).filter(
                UbicacionCliente.id_ubicacion == cliente_data['id_ubicacion_principal']
            ).first()
            
            if not ubicacion:
                raise HTTPException(
                    status_code=400, 
                    detail="La ubicación principal especificada no existe"
                )
            
            # Verificar que la ubicación pertenezca al cliente
            if ubicacion.cod_cliente != cliente_data.get('cod_cliente'):
                raise HTTPException(
                    status_code=400, 
                    detail="La ubicación principal debe pertenecer al cliente"
                )
        
        nuevo_cliente = Cliente(**cliente_data)
        db.add(nuevo_cliente)
        db.commit()
        db.refresh(nuevo_cliente)
        return nuevo_cliente
        
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error al crear cliente: {str(e)}")

def update_cliente(db: Session, cod_cliente: str, cliente_data: dict):
    """
    Actualiza un cliente con validación de ubicación principal
    """
    try:
        cliente = db.query(Cliente).filter(Cliente.cod_cliente == cod_cliente).first()
        if not cliente:
            raise HTTPException(status_code=404, detail="Cliente no encontrado")
        
        # Validar ubicación principal si se proporciona
        if cliente_data.get('id_ubicacion_principal'):
            ubicacion = db.query(UbicacionCliente).filter(
                UbicacionCliente.id_ubicacion == cliente_data['id_ubicacion_principal']
            ).first()
            
            if not ubicacion:
                raise HTTPException(
                    status_code=400, 
                    detail="La ubicación principal especificada no existe"
                )
            
            # Verificar que la ubicación pertenezca al cliente
            if ubicacion.cod_cliente != cod_cliente:
                raise HTTPException(
                    status_code=400, 
                    detail="La ubicación principal debe pertenecer al cliente"
                )
        
        # Actualizar campos
        for key, value in cliente_data.items():
            setattr(cliente, key, value)
        
        db.commit()
        db.refresh(cliente)
        return cliente
        
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error al actualizar cliente: {str(e)}")

def delete_cliente(db: Session, cod_cliente: str):
    """
    Elimina un cliente y maneja las relaciones con ubicaciones
    """
    try:
        cliente = db.query(Cliente).filter(Cliente.cod_cliente == cod_cliente).first()
        if not cliente:
            raise HTTPException(status_code=404, detail="Cliente no encontrado")
        
        # Verificar si tiene ubicaciones asociadas
        ubicaciones_count = db.query(UbicacionCliente).filter(
            UbicacionCliente.cod_cliente == cod_cliente
        ).count()
        
        if ubicaciones_count > 0:
            raise HTTPException(
                status_code=400, 
                detail=f"No se puede eliminar el cliente. Tiene {ubicaciones_count} ubicaciones asociadas. Elimine primero las ubicaciones."
            )
        
        db.delete(cliente)
        db.commit()
        return {"mensaje": "Cliente eliminado correctamente"}
        
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error al eliminar cliente: {str(e)}")

def get_clientes_con_ubicaciones(db: Session):
    """
    Obtiene todos los clientes con información detallada de sus ubicaciones
    """
    clientes = db.query(Cliente).options(
        joinedload(Cliente.ubicaciones)
    ).all()
    
    result = []
    for cliente in clientes:
        cliente_dict = {
            "cod_cliente": cliente.cod_cliente,
            "identificacion": cliente.identificacion,
            "nombre": cliente.nombre,
            "direccion": cliente.direccion,
            "celular": cliente.celular,
            "correo": cliente.correo,
            "tipo_cliente": cliente.tipo_cliente,
            "razon_social": cliente.razon_social,
            "sector": cliente.sector,
            "fecha_registro": cliente.fecha_registro,
            "id_ubicacion_principal": cliente.id_ubicacion_principal,
            "ubicaciones": []
        }
        
        # Agregar información de ubicaciones
        for ubicacion in cliente.ubicaciones:
            ubicacion_dict = {
                "id_ubicacion": ubicacion.id_ubicacion,
                "latitud": float(ubicacion.latitud),
                "longitud": float(ubicacion.longitud),
                "direccion": ubicacion.direccion,
                "sector": ubicacion.sector,
                "referencia": ubicacion.referencia,
                "fecha_registro": ubicacion.fecha_registro
            }
            cliente_dict["ubicaciones"].append(ubicacion_dict)
        
        result.append(cliente_dict)
    
    return result

def export_clientes_to_excel(db: Session):
    """
    Exporta los clientes a un archivo Excel con información de ubicación principal
    """
    try:
        clientes = db.query(Cliente).options(
            joinedload(Cliente.ubicaciones)
        ).all()
        
        if not clientes:
            raise HTTPException(status_code=404, detail="No hay clientes para exportar")
        
        # Preparar datos para el DataFrame
        data = []
        for cliente in clientes:
            # Formatear fechas - manejar valores None
            fecha_registro_str = ""
            if cliente.fecha_registro:
                try:
                    fecha_registro_str = cliente.fecha_registro.strftime("%d/%m/%Y")
                except Exception as e:
                    print(f"Error formateando fecha para cliente {cliente.cod_cliente}: {e}")
                    fecha_registro_str = str(cliente.fecha_registro)
            
            # Obtener información de ubicación principal
            ubicacion_principal_info = "Sin ubicación"
            if cliente.id_ubicacion_principal:
                ubicacion_principal = next(
                    (ub for ub in cliente.ubicaciones if ub.id_ubicacion == cliente.id_ubicacion_principal), 
                    None
                )
                if ubicacion_principal:
                    ubicacion_principal_info = f"{ubicacion_principal.sector} - {ubicacion_principal.direccion[:50]}..."
            
            data.append({
                'Código Cliente': cliente.cod_cliente or '',
                'Identificación': cliente.identificacion or '',
                'Nombre': cliente.nombre or '',
                'Dirección': cliente.direccion or '',
                'Celular': cliente.celular or '',
                'Correo': cliente.correo or '',
                'Tipo Cliente': cliente.tipo_cliente or '',
                'Razón Social': cliente.razon_social or '',
                'Sector': cliente.sector or '',
                'Ubicación Principal': ubicacion_principal_info,
                'Total Ubicaciones': len(cliente.ubicaciones),
                'Fecha Registro': fecha_registro_str
            })
        
        # Crear DataFrame
        df = pd.DataFrame(data)
        
        # Crear archivo Excel con diseño profesional
        wb = Workbook()
        ws = wb.active
        ws.title = "Lista de Clientes"
        
        # Configurar estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Estilo para las celdas de datos
        data_alignment = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Agregar título principal
        ws.merge_cells('A1:L1')  # Actualizado para incluir las nuevas columnas
        title_cell = ws['A1']
        title_cell.value = f"LISTA DE CLIENTES CON UBICACIONES - {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Agregar una fila vacía
        ws.append([])
        
        # Agregar encabezados
        headers = list(df.columns)
        ws.append(headers)
        
        # Aplicar estilo a los encabezados
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Agregar datos
        for row_num, row_data in enumerate(df.values, 4):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = data_alignment
                cell.border = thin_border
                
                # Aplicar color alternado a las filas
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Ajustar ancho de columnas
        column_widths = {
            'A': 15,  # Código Cliente
            'B': 15,  # Identificación
            'C': 25,  # Nombre
            'D': 30,  # Dirección
            'E': 15,  # Celular
            'F': 25,  # Correo
            'G': 15,  # Tipo Cliente
            'H': 25,  # Razón Social
            'I': 20,  # Sector
            'J': 35,  # Ubicación Principal
            'K': 15,  # Total Ubicaciones
            'L': 15   # Fecha Registro
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Agregar información de resumen al final
        last_row = len(df) + 5
        ws.cell(row=last_row, column=1, value="Total de clientes:").font = Font(bold=True)
        ws.cell(row=last_row, column=2, value=len(df)).font = Font(bold=True)
        
        ws.cell(row=last_row + 1, column=1, value="Clientes con ubicación principal:").font = Font(bold=True)
        clientes_con_ubicacion = len([c for c in clientes if c.id_ubicacion_principal])
        ws.cell(row=last_row + 1, column=2, value=clientes_con_ubicacion).font = Font(bold=True)
        
        # Crear el archivo en memoria
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Retornar los bytes del archivo
        return excel_buffer.getvalue()
        
    except Exception as e:
        print(f"Error detallado al exportar clientes a Excel: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error al generar archivo Excel: {str(e)}")